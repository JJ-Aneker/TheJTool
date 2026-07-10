"""
Parser de cuestionarios de seguridad IT en Excel.

Objetivo: leer un .xlsx (formato desconocido de antemano, cada cliente lo
estructura distinto) y devolver una lista normalizada de preguntas, cada
una con: hoja, categoría/sección, texto de la pregunta, respuesta existente
(si la hay), fila/columna de origen (para poder rellenar después) y el
método de detección usado (para poder revisar los casos "heurísticos" con
más cuidado).

Estrategia en dos fases:
  1. Intentar detectar una fila de cabeceras real (Pregunta/Respuesta/...)
     en las primeras N filas de cada hoja. Si se encuentra, se usa mapeo
     de columnas por cabecera.
  2. Si no hay cabeceras claras, modo heurístico: se buscan celdas que
     parezcan preguntas (terminan en "?", o contienen palabras clave tipo
     "indicar", "describir") y se intenta capturar la celda de respuesta
     como la celda vacía más cercana a la derecha o debajo.

Las celdas combinadas de una sola columna ancha (p.ej. "1. AUTENTICACIÓN")
se tratan como cabeceras de sección y se propagan a las preguntas
siguientes hasta la próxima sección.
"""

from dataclasses import dataclass, field
from openpyxl import load_workbook
import re

# Palabras clave para detectar columnas por cabecera (case-insensitive).
# Nota: se evitan palabras demasiado genéricas (p.ej. "control" a secas)
# porque colisionan con títulos de sección tipo "Access Control".
HEADER_KEYWORDS = {
    "question": ["pregunta", "question", "requisito", "requirement", "item"],
    "answer": ["respuesta", "answer", "response"],
    "evidence": ["evidencia", "evidence", "comentario", "comment", "notes", "nota", "observaciones"],
    "category": ["categoría", "categoria", "category", "sección", "seccion", "section"],
    "id": ["id", "ref.", "referencia", "reference", "no.", "n°"],
}
MIN_ROLES_FOR_HEADER = 2  # una fila necesita al menos 2 roles distintos para contar como cabecera

# Señales de que una celda es una pregunta en modo heurístico
QUESTION_MARKERS = ["?", "indicar", "describir", "especificar", "detallar", "explicar"]

MAX_HEADER_SEARCH_ROWS = 10
MIN_AVG_LEN_FOR_INFERRED_QUESTION = 25
INFERENCE_SAMPLE_ROWS = 15
MIN_QUESTION_TEXT_LENGTH = 5  # descarta ruido tipo "?" suelto

# Hojas conocidas como tablas de referencia/mapeo interno (no son
# preguntas a responder por el proveedor). Confirmado con plantilla ING
# "Third Party IT Security Compliance". Ajustar por cliente/plantilla si
# hace falta pasando excluded_sheets a parse_workbook().
DEFAULT_EXCLUDED_SHEETS = {
    "mapper", "mapperold", "itss v3 mapper", "rating",
    "process catalogue", "itss",
}


def _infer_question_column(sheet, header_row, col_map, merged_lookup):
    """
    Cuando se detectó cabecera (p.ej. 'answers') pero ninguna columna se
    reconoció como 'question' por etiqueta, se infiere por contenido:
    la columna sin rol asignado con mayor longitud media de texto en las
    filas siguientes suele ser la de la pregunta (texto largo) frente a
    columnas de categoría/ID/referencia (texto corto).
    """
    used_cols = set(col_map.values())
    max_col = sheet.max_column
    sample_end = min(header_row + INFERENCE_SAMPLE_ROWS, sheet.max_row)

    best_col, best_avg = None, 0
    for col_idx in range(1, max_col + 1):
        if col_idx in used_cols:
            continue
        lengths = []
        for row_idx in range(header_row + 1, sample_end + 1):
            val = _cell_value(sheet, row_idx, col_idx, merged_lookup)
            if val:
                lengths.append(len(str(val)))
        if lengths:
            avg = sum(lengths) / len(lengths)
            if avg > best_avg:
                best_avg, best_col = avg, col_idx

    if best_col and best_avg >= MIN_AVG_LEN_FOR_INFERRED_QUESTION:
        return best_col
    return None


@dataclass
class Question:
    sheet: str
    section: str
    question_id: str
    text: str
    existing_answer: str
    evidence_note: str
    cell_ref: str            # celda de la pregunta, p.ej. "C5"
    answer_cell_ref: str     # celda donde debería ir la respuesta (para rellenar luego)
    detection_method: str    # "header" o "heuristic" — útil para revisión humana
    confidence: str          # "alta" o "baja" (heurístico siempre es "baja")


def _norm(text):
    """Normaliza texto de celda para comparar cabeceras."""
    if text is None:
        return ""
    return str(text).strip().lower()


def _cell_value(sheet, row, col, merged_lookup):
    """
    Devuelve el valor de una celda, resolviendo el caso de que pertenezca
    a un rango combinado (las celdas combinadas solo tienen valor en la
    esquina superior izquierda; el resto son None en openpyxl).
    """
    key = (row, col)
    if key in merged_lookup:
        row, col = merged_lookup[key]
    val = sheet.cell(row=row, column=col).value
    return val


def _build_merged_lookup(sheet):
    """
    Construye un diccionario {(fila, col): (fila_origen, col_origen)} para
    poder resolver el valor real de cualquier celda dentro de un rango
    combinado.
    """
    lookup = {}
    for merged_range in sheet.merged_cells.ranges:
        top_left = (merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, col)] = top_left
    return lookup


def _find_section_rows(sheet):
    """
    Identifica filas que son títulos de sección: rangos combinados de una
    sola fila que abarcan 2+ columnas (p.ej. "1. AUTENTICACIÓN" en A1:D1).
    Devuelve {fila: texto_seccion}.
    """
    section_rows = {}
    for merged_range in sheet.merged_cells.ranges:
        spans_one_row = merged_range.min_row == merged_range.max_row
        spans_multiple_cols = merged_range.max_col > merged_range.min_col
        if spans_one_row and spans_multiple_cols:
            text = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
            if text:
                section_rows[merged_range.min_row] = str(text).strip()
    return section_rows


def _detect_header_row(sheet, merged_lookup, section_rows):
    """
    Busca en las primeras filas una que contenga cabeceras reconocibles.
    Exige al menos MIN_ROLES_FOR_HEADER roles distintos para evitar que
    un título de sección con una palabra suelta ("...Access Control...")
    se confunda con una cabecera real. Las filas de sección ya detectadas
    se descartan directamente.
    Devuelve (fila_index, mapeo_columnas) o (None, None).
    """
    max_row = min(MAX_HEADER_SEARCH_ROWS, sheet.max_row)
    max_col = sheet.max_column

    for row_idx in range(1, max_row + 1):
        if row_idx in section_rows:
            continue

        col_map = {}
        for col_idx in range(1, max_col + 1):
            val = _norm(_cell_value(sheet, row_idx, col_idx, merged_lookup))
            if not val:
                continue
            for role, keywords in HEADER_KEYWORDS.items():
                if role in col_map:
                    continue
                if any(re.search(r"\b" + re.escape(kw) + r"\b", val) for kw in keywords):
                    col_map[role] = col_idx

        # Caso normal: cabecera con "question" etiquetado explícitamente
        if "question" in col_map and len(col_map) >= MIN_ROLES_FOR_HEADER:
            header_texts = {
                role: _norm(_cell_value(sheet, row_idx, col, merged_lookup))
                for role, col in col_map.items()
            }
            return row_idx, col_map, header_texts

        # Caso "Cyber Security": hay cabecera real (p.ej. solo 'answers')
        # pero ninguna columna se etiquetó como pregunta. Se infiere por
        # contenido: la columna sin rol con mayor longitud media de texto.
        if col_map and "question" not in col_map:
            inferred_col = _infer_question_column(sheet, row_idx, col_map, merged_lookup)
            if inferred_col:
                col_map = {**col_map, "question": inferred_col}
                header_texts = {
                    role: _norm(_cell_value(sheet, row_idx, col, merged_lookup))
                    for role, col in col_map.items()
                }
                return row_idx, col_map, header_texts

    return None, None, {}


def _extract_with_headers(sheet, sheet_name, header_row, col_map, merged_lookup, section_rows, header_texts):
    """Extrae preguntas cuando se detectó una fila de cabeceras clara."""
    questions = []
    q_col = col_map["question"]

    # Si había un título de sección ANTES de la fila de cabecera, lo
    # heredamos como sección inicial (caso típico: título de bloque,
    # luego la cabecera de columnas, luego las preguntas).
    current_section = ""
    for row_idx in sorted(section_rows):
        if row_idx < header_row:
            current_section = section_rows[row_idx]
        else:
            break

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        if row_idx in section_rows:
            current_section = section_rows[row_idx]
            continue

        question_text = _cell_value(sheet, row_idx, q_col, merged_lookup)
        if question_text in (None, "") or not str(question_text).strip():
            continue

        # Cabecera repetida dentro de la misma hoja (una por sub-sección):
        # si el texto coincide con la cabecera original, se salta.
        if _norm(question_text) == header_texts.get("question", "__no_match__"):
            continue

        question_text = str(question_text).strip()
        if len(question_text) < MIN_QUESTION_TEXT_LENGTH:
            continue

        answer = _cell_value(sheet, row_idx, col_map["answer"], merged_lookup) if "answer" in col_map else None
        evidence = _cell_value(sheet, row_idx, col_map["evidence"], merged_lookup) if "evidence" in col_map else None
        qid = _cell_value(sheet, row_idx, col_map["id"], merged_lookup) if "id" in col_map else None

        # Si hay columna de categoría explícita, tiene prioridad sobre la
        # sección heredada de un título combinado (es más específica).
        section = current_section
        if "category" in col_map:
            cat_val = _cell_value(sheet, row_idx, col_map["category"], merged_lookup)
            if cat_val not in (None, ""):
                section = str(cat_val).strip()

        answer_col = col_map.get("answer", q_col + 1)

        questions.append(Question(
            sheet=sheet_name,
            section=section,
            question_id=str(qid) if qid is not None else "",
            text=question_text,
            existing_answer=str(answer).strip() if answer not in (None, "") else "",
            evidence_note=str(evidence).strip() if evidence not in (None, "") else "",
            cell_ref=sheet.cell(row=row_idx, column=q_col).coordinate,
            answer_cell_ref=sheet.cell(row=row_idx, column=answer_col).coordinate,
            detection_method="header",
            confidence="alta",
        ))

    return questions


def _extract_heuristic(sheet, sheet_name, merged_lookup, section_rows):
    """
    Modo heurístico: sin cabeceras claras. Busca celdas que parezcan
    preguntas y asume que la respuesta va en la celda vacía inmediatamente
    a la derecha, o si no, en la celda vacía inmediatamente debajo.
    Marca todo como confianza baja para forzar revisión humana.
    """
    questions = []
    current_section = ""

    for row_idx in range(1, sheet.max_row + 1):
        if row_idx in section_rows:
            current_section = section_rows[row_idx]
            continue

        for col_idx in range(1, sheet.max_column + 1):
            val = _cell_value(sheet, row_idx, col_idx, merged_lookup)
            if val is None:
                continue
            text = str(val).strip()
            if not text:
                continue

            is_question = any(marker in text.lower() for marker in QUESTION_MARKERS)
            if not is_question:
                # ¿Es un título de sección? (negrita+cursiva no accesible aquí sin
                # inspeccionar estilo; usamos heurística de mayúsculas/longitud corta)
                if text.isupper() and len(text) < 80:
                    current_section = text
                continue

            if len(text) < MIN_QUESTION_TEXT_LENGTH:
                continue

            # Buscar celda de respuesta: primero a la derecha, luego debajo
            answer_cell = sheet.cell(row=row_idx, column=col_idx + 1)
            answer_ref = answer_cell.coordinate
            if answer_cell.value not in (None, ""):
                # la celda a la derecha ya tiene contenido -> probar debajo
                below_cell = sheet.cell(row=row_idx + 1, column=col_idx)
                if below_cell.value in (None, ""):
                    answer_ref = below_cell.coordinate

            questions.append(Question(
                sheet=sheet_name,
                section=current_section,
                question_id="",
                text=text,
                existing_answer="",
                evidence_note="",
                cell_ref=sheet.cell(row=row_idx, column=col_idx).coordinate,
                answer_cell_ref=answer_ref,
                detection_method="heuristic",
                confidence="baja",
            ))

    return questions


def parse_workbook(path, excluded_sheets=None):
    """
    Punto de entrada principal. Devuelve una lista de objetos Question
    de todas las hojas del fichero.

    excluded_sheets: nombres de hoja (case-insensitive) a ignorar por
    completo, p.ej. tablas de referencia/mapeo interno que no son
    preguntas a responder. Si no se indica, se usa DEFAULT_EXCLUDED_SHEETS.
    """
    excluded = {s.lower() for s in (excluded_sheets if excluded_sheets is not None else DEFAULT_EXCLUDED_SHEETS)}

    wb = load_workbook(path, data_only=True)
    all_questions = []

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in excluded:
            continue

        sheet = wb[sheet_name]
        merged_lookup = _build_merged_lookup(sheet)
        section_rows = _find_section_rows(sheet)

        header_row, col_map, header_texts = _detect_header_row(sheet, merged_lookup, section_rows)

        if header_row is not None:
            qs = _extract_with_headers(sheet, sheet_name, header_row, col_map, merged_lookup, section_rows, header_texts)
        else:
            qs = _extract_heuristic(sheet, sheet_name, merged_lookup, section_rows)

        all_questions.extend(qs)

    return all_questions


if __name__ == "__main__":
    import sys
    import json

    path = sys.argv[1] if len(sys.argv) > 1 else "cuestionario_prueba.xlsx"
    results = parse_workbook(path)

    print(f"\n{len(results)} preguntas extraídas de '{path}'\n")
    for q in results:
        print(f"[{q.sheet}] ({q.detection_method}, confianza {q.confidence}) {q.cell_ref}")
        if q.section:
            print(f"  Sección: {q.section}")
        if q.question_id:
            print(f"  ID: {q.question_id}")
        print(f"  Pregunta: {q.text}")
        if q.existing_answer:
            print(f"  Respuesta existente: {q.existing_answer}")
        print(f"  Celda destino respuesta: {q.answer_cell_ref}")
        print()

    # Export JSON también, para que otros módulos (RAG, generación) lo consuman
    out_path = path.rsplit(".", 1)[0] + "_preguntas.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump([q.__dict__ for q in results], f, ensure_ascii=False, indent=2)
    print(f"JSON exportado a: {out_path}")
